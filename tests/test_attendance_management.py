"""Integration tests for the Attendance Management features (details, filters, exports, stats)."""

import csv
import io
from openpyxl import load_workbook
from app.extensions import db
from app.models import Person, AttendanceRecord, Setting, utcnow
from tests.conftest import login, observation, unit_vector
from tests.test_api import enroll, open_session


def test_session_detail_and_filtering(client, fake_vision):
    login(client)

    # 1. Create students
    now = utcnow()
    p1 = Person(code="R01", full_name="Student One", department="CS", year=1, enrolled_at=now)
    p2 = Person(code="R02", full_name="Student Two", department="ECE", year=2, enrolled_at=now)
    db.session.add_all([p1, p2])
    db.session.commit()

    # Enroll
    from app.models import FaceEmbedding
    from app.services.face_engine import MODEL_NAME
    emb1 = FaceEmbedding(person=p1, is_centroid=True, quality_score=0.9, model_name=MODEL_NAME)
    emb1.set_vector(unit_vector(axis=0))
    emb2 = FaceEmbedding(person=p2, is_centroid=True, quality_score=0.9, model_name=MODEL_NAME)
    emb2.set_vector(unit_vector(axis=1))
    db.session.add_all([emb1, emb2])
    db.session.commit()

    from app.services import matcher
    matcher.invalidate_gallery()

    session_id = open_session(client)

    # Mark p1 present, p2 remains absent
    fake_vision([[observation(unit_vector(axis=0))]])
    client.post("/api/v1/recognize", json={"image": "aW1n", "session_id": session_id})

    # 2. Get session details page
    response = client.get(f"/sessions/{session_id}")
    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "Student One" in html
    assert "Student Two" in html
    assert "present" in html
    assert "absent" in html

    # 3. Filter by status=present
    response = client.get(f"/sessions/{session_id}?status=present")
    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "Student One" in html
    assert "Student Two" not in html

    # 4. Filter by status=absent
    response = client.get(f"/sessions/{session_id}?status=absent")
    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "Student One" not in html
    assert "Student Two" in html

    # 5. Search query
    response = client.get(f"/sessions/{session_id}?q=Student+One")
    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "Student One" in html
    assert "Student Two" not in html


def test_session_exports(client, fake_vision):
    login(client)

    p1 = Person(code="R03", full_name="Student Three", department="CS", year=3, enrolled_at=utcnow())
    db.session.add(p1)
    db.session.commit()

    # Centroid
    from app.models import FaceEmbedding
    from app.services.face_engine import MODEL_NAME
    emb = FaceEmbedding(person=p1, is_centroid=True, quality_score=0.9, model_name=MODEL_NAME)
    emb.set_vector(unit_vector(axis=0))
    db.session.add(emb)
    db.session.commit()

    from app.services import matcher
    matcher.invalidate_gallery()

    session_id = open_session(client)

    # 1. Export CSV
    response = client.get(f"/sessions/{session_id}/export?format=csv")
    assert response.status_code == 200
    assert "text/csv" in response.headers["Content-Type"]
    assert f"attendance_session_{session_id}.csv" in response.headers["Content-Disposition"]

    csv_data = response.get_data(as_text=True)
    # Excel UTF-8 BOM should be prepended
    assert csv_data.startswith("\ufeff")
    reader = csv.reader(io.StringIO(csv_data.lstrip("\ufeff")))
    rows = list(reader)
    assert rows[0] == ["Roll Number", "Name", "Department", "Year", "Email", "Status", "Marked At", "Confidence"]
    assert rows[1][0] == "R03"
    assert rows[1][1] == "Student Three"
    assert rows[1][5] == "Absent"

    # 2. Export Excel (.xlsx binary)
    response = client.get(f"/sessions/{session_id}/export?format=excel")
    assert response.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["Content-Type"]
    assert f"attendance_session_{session_id}.xlsx" in response.headers["Content-Disposition"]

    # Read binary bytes using openpyxl
    wb = load_workbook(io.BytesIO(response.get_data()))
    sheet = wb.active
    assert sheet.title == "Attendance"
    assert sheet.cell(row=1, column=1).value == "Roll Number"
    assert sheet.cell(row=2, column=1).value == "R03"
    assert sheet.cell(row=2, column=2).value == "Student Three"
    assert sheet.cell(row=2, column=6).value == "Absent"


def test_attendance_statistics_and_summary_export(client):
    login(client)

    p1 = Person(code="R04", full_name="Student Four", department="ECE", year=4, enrolled_at=utcnow())
    db.session.add(p1)
    db.session.commit()

    # 1. View statistics page
    response = client.get("/attendance/stats")
    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "Reports" in html
    assert "Student Four" in html

    # 2. Export overall summary reports
    response = client.get("/attendance/stats/export?format=csv")
    assert response.status_code == 200
    assert "text/csv" in response.headers["Content-Type"]
    
    response = client.get("/attendance/stats/export?format=excel")
    assert response.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["Content-Type"]


def test_attendance_cooldown_enforcement(client, fake_vision):
    login(client)

    # 1. Enable cooldown of 5 minutes
    Setting.set("attendance_cooldown_minutes", "5")
    db.session.commit()

    # 2. Create and enroll student
    p = Person(code="C001", full_name="Cooldown Test Student", enrolled_at=utcnow())
    db.session.add(p)
    db.session.commit()

    from app.models import FaceEmbedding
    from app.services.face_engine import MODEL_NAME
    emb = FaceEmbedding(person=p, is_centroid=True, quality_score=0.9, model_name=MODEL_NAME)
    emb.set_vector(unit_vector(axis=0))
    db.session.add(emb)
    db.session.commit()

    from app.services import matcher
    matcher.invalidate_gallery()

    # 3. Create two sessions
    session_id_1 = open_session(client, name="Session One")
    session_id_2 = open_session(client, name="Session Two")

    # 4. Mark in Session One
    fake_vision([[observation(unit_vector(axis=0))]])
    res1 = client.post("/api/v1/recognize", json={"image": "aW1n", "session_id": session_id_1})
    assert res1.status_code == 200
    body1 = res1.get_json()
    assert body1["attendance"]["created"] is True
    assert body1["attendance"]["status"] == "present"

    # 5. Try to mark in Session Two immediately (should be blocked by cooldown)
    fake_vision([[observation(unit_vector(axis=0))]])
    res2 = client.post("/api/v1/recognize", json={"image": "aW1n", "session_id": session_id_2})
    assert res2.status_code == 200
    body2 = res2.get_json()
    assert body2["attendance"]["created"] is False

    # 6. Disable cooldown
    Setting.set("attendance_cooldown_minutes", "0")
    db.session.commit()

    # 7. Try to mark in Session Two again (should succeed)
    fake_vision([[observation(unit_vector(axis=0))]])
    res3 = client.post("/api/v1/recognize", json={"image": "aW1n", "session_id": session_id_2})
    assert res3.status_code == 200
    body3 = res3.get_json()
    assert body3["attendance"]["created"] is True


def test_inactive_student_attendance_in_details_and_exports(client, fake_vision):
    login(client)

    # 1. Create and enroll student
    p = Person(code="I001", full_name="Inactive History Student", enrolled_at=utcnow())
    db.session.add(p)
    db.session.commit()

    from app.models import FaceEmbedding
    from app.services.face_engine import MODEL_NAME
    emb = FaceEmbedding(person=p, is_centroid=True, quality_score=0.9, model_name=MODEL_NAME)
    emb.set_vector(unit_vector(axis=0))
    db.session.add(emb)
    db.session.commit()

    from app.services import matcher
    matcher.invalidate_gallery()

    session_id = open_session(client, name="History Session")

    # 2. Mark present
    fake_vision([[observation(unit_vector(axis=0))]])
    client.post("/api/v1/recognize", json={"image": "aW1n", "session_id": session_id})

    # 3. Deactivate student
    p.is_active = False
    db.session.commit()
    matcher.invalidate_gallery()

    # 4. View session details page
    response = client.get(f"/sessions/{session_id}")
    assert response.status_code == 200
    html = response.get_data(as_text=True)
    # The student and their present status must still be visible in the history
    assert "Inactive History Student" in html
    assert "present" in html

    # 5. Export CSV
    response = client.get(f"/sessions/{session_id}/export?format=csv")
    assert response.status_code == 200
    csv_data = response.get_data(as_text=True)
    assert "Inactive History Student" in csv_data

    # 6. Export Excel
    response = client.get(f"/sessions/{session_id}/export?format=excel")
    assert response.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["Content-Type"]
